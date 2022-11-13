import openpyxl
import os
import requests
from datetime import datetime
import time
import warnings
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
# from pyvirtualdisplay import Display
import subprocess
from lxml import html
from bs4 import BeautifulSoup as BS
import pandas as pd
import time
import mysql.connector
# import MySQLdb


class CrawlLinkSpider():
    def __init__(self):
        # self.conn = mysql.connector.connect(host="75.126.130.244", user="crawler@web", passwd="data@server", db="dell_pl_crawl_latest",charset="utf8", use_unicode=True)
        self.conn = mysql.connector.connect(host="localhost", user="root", passwd="nixis123", db="crawling",charset="utf8", use_unicode=True)
        self.cursor = self.conn.cursor()
        self.output_file_name = "PL_Crawl_Mediamarkt_DE_Dell_GPT_Laptop_Output.xlsx"
        self.driver = ""
        self.output = pd.DataFrame()

    def Update_IP(self):
        
        self.driver.switch_to.window(self.driver.window_handles[1])
        try:
            country = self.driver.find_element_by_id("connected-country").text
        except:
            country = "Germany"
        
        self.used_country.append(country)
        print(self.used_country)
        try:
            self.driver.find_element_by_id("connected-disconnect-button").click()
        except:
            pass
        time.sleep(2)
        print(".....")
        print(self.driver.find_elements_by_class_name('server-item__server__server-info__server-label'))
        for each in self.driver.find_elements_by_class_name('server-item__server__server-info__server-label'):
            if(each.text not in self.used_country):
                print(each.text)
                each.click()
                break
        conn_text="No"
        while conn_text=="No":
            try:
                conn_text = self.driver.find_element_by_id("connected-disconnect-button").text
            except:
                conn_text="No"            
        # time.sleep(4)
        self.driver.switch_to.window(self.driver.window_handles[0])
        self.driver.refresh()
        soup = BS(self.driver.page_source,'html.parser')
        if(soup.title.text == "Saturn" or soup.title.text == "MediaMarkt"):
            self.Update_IP()

    def crawl_link(self,driver,url,cat,landing_page):
        driver.get(url)
        tree = html.fromstring(driver.page_source)
        soup = BS(driver.page_source,'html.parser')
        # soup = BS(self.driver.page_source,"html.parser")
        if(soup.title.text == "Saturn" or soup.title.text == "MediaMarkt"):
            self.Update_IP()        
        # links = tree.xpath("//a[@class='Linkstyled__StyledLinkRouter-sc-1drhx1h-2 OFLqH ProductListItemstyled__StyledLink-sc-1627gef-0 cgpKZO']/@href")    
        # titles = soup.find_all("p","Typostyled__StyledInfoTypo-sc-1jga2g7-0 clXCSd")

        links = tree.xpath("//a[@class='Linkstyled__StyledLinkRouter-sc-1drhx1h-2 hihJjl ProductListItemstyled__StyledLink-sc-16qx04k-0 dYJAjV']/@href")
        titles = soup.find_all("p","Typostyled__StyledInfoTypo-sc-1jga2g7-0 iLnJNN")

        for link,title in zip(links,titles):
            c = {
                # "Script_Name":"PL_Crawl_Mediamarkt_Germany_Dell_GPT_{}".format(cat),
                "Date":time.strftime("%Y-%m-%d"),
                "Time":time.strftime("%H:%M:%S"),
                "Landing_Page_url":landing_page,
                "Product_title":title.text,
                "product_link":"https://www.mediamarkt.de"+link,
                "product_id" : link.replace(".html","").split("-")[-1],
                "Category":cat,
                # "date":time.strftime("%Y-%m-%d"),
                # "time":time.strftime("%H:%M:%S"),

            }
            print(c)
            query  = "insert into mediamarkt_de(product_title,product_link,category,date,landing_page_url)values(%s,%s,%s,%s,%s)"
            values = (c["Product_title"],c["product_link"],c["Category"],c["Date"],c["Landing_Page_url"])
            self.cursor.execute(query, values)
            self.conn.commit()

            # df = pd.DataFrame([c])
            # cursor = self.conn.cursor()
            # query = "insert into mediamarket_de(SCRIPT_NAME,DATE,TIME,LANDING_PAGE_URL,PRODUCT_TITLE,PRODUCT_LINK,PRODUCT_CODE,CUSTOMIZED_URL,NEW_URL,DEACTIVATED,REACTIVATED,LAST_DEACTIVATED,LAST_REACTIVATED)VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
            # values = (c["Script_Name"],c["Date"],c["Time"],c["Landing_Page_url"],c["Product_title"],c["product_link"],c["product_id"],c["Customized_url"],c["New_url"],c["Deactivated"],c["Reactivated"],c["last_deactivated"],c["last_reactivated"])
            # cursor.execute(query, values)
            # self.conn.commit()
            # cursor.execute("select * from pl_master where script_name = '{}' and product_code='{}' and deactivated=0".format(c["Script_Name"],c["product_id"]))
            # if(not cursor.fetchone()):
            #     query = "insert into saturn_de(SCRIPT_NAME,DATE,TIME,LANDING_PAGE_URL,PRODUCT_TITLE,PRODUCT_LINK,PRODUCT_CODE,CUSTOMIZED_URL,NEW_URL,DEACTIVATED,REACTIVATED,LAST_DEACTIVATED,LAST_REACTIVATED)VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
            #     values = (c["Script_Name"],c["Date"],c["Time"],c["Landing_Page_url"],c["Product_title"],c["product_link"],c["product_id"],c["Customized_url"],c["New_url"],c["Deactivated"],c["Reactivated"],c["last_deactivated"],c["last_reactivated"])
            #     cursor.execute(query, values)
            #     self.conn.commit()
            # self.output = self.output.append(df,ignore_index=True)
        
        # load_more = soup.find("button",attrs={"data-test":"mms-search-srp-loadmore"})
        load_more = soup.find("button",attrs={"data-test":"mms-search-srp-loadmore"})
        if(load_more):
            print("loadmore.......")
            if(driver.current_url.find("page") == -1):
                next_page_url = driver.current_url+"&page=2"
            else:
                page_num = driver.current_url.split("page=")[1]
                next_page_url = driver.current_url.replace("page={}".format(page_num),"page={}".format(int(page_num)+1))

            self.crawl_link(driver,next_page_url,cat,landing_page)
        # else:
        #     self.output.to_excel(self.output_file_name,index=False)

    def close_browser(self):
        #for i in self.driver.window_handles:
        #    self.driver.switch_to.window(i)
        #    self.driver.close()

        self.driver.switch_to.window(self.driver.window_handles[1])
        self.driver.find_element_by_id("connected-disconnect-button").click()
        time.sleep(2)
        self.driver.find_element_by_id("logout-button").click()
        #self.driver.close()
        #self.driver.switch_to.window(self.driver.window_handles[0])
        #self.driver.quit()
        for i in self.driver.window_handles:
            self.driver.switch_to.window(i)
            self.driver.close()		
        print("process completed")


    def open_script(self,driver):
        try:
            driver.execute_script("window.open('chrome-extension://oofgbpoabipfcfjapgnbbjjaenockbdp/popup.html')")
        except:
            self.open_script(driver)

    def Resolve_Captcha(self):
        print("Process Started")
        cwd = os.getcwd()
        #os.system("C:/Windows/System32/taskkill /im chrome.exe /f")
        # os.system("taskkill /im chrome.exe /f")
        # os.system('cmd /k "dir"')
        # subprocess.call('start cd c:\\',shell=True)
        warnings.filterwarnings("ignore", category=DeprecationWarning)
        chrome_options = Options()
        # chrome_options.add_argument('--user-agent="Mozilla/5.0 (Windows NT 6.1;WOW64; rv:50.0) Gecko/20100101 Firefox/50.0"')


        # chrome_options.add_argument("user-data-dir=C:\\Users\\sunil.dhasmana\\AppData\\Local\\Google\\Chrome\\User Data\\Default")
        # chrome_options.add_experimental_option('debuggerAddress', 'localhost:9250')
        # proxy = 'http://scraperapi.country_code=us:a6438ab03fee3e0af7053fbbcaa5c20c@proxy-server.scraperapi.com:8001'
        # chrome_options.add_argument('--proxy-server=%s' % proxy)
        # chrome_options.add_extension(r'' + cwd + "\\1.2.7_0.crx")

        # from selenium.webdriver.common.proxy import Proxy

        # capabilities = webdriver.DesiredCapabilities.CHROME
        excel_input = cwd + "/Input_file.xlsx"
        wb = openpyxl.load_workbook(excel_input)
        sht_login = wb.get_sheet_by_name('Login Info')
        sht_input = wb.get_sheet_by_name('Input')
        sht_output = wb.get_sheet_by_name('Output')
        username = sht_login.cell(1,2).value
        password = sht_login.cell(2,2).value
        chrmprof = sht_login.cell(3,2).value
        # subprocess.call('start chrome.exe -remote-debugging-port=9014 --user-data-dir-' + chrmprof,shell=True)
        # chrome_options.add_experimental_option("debuggerAddress","localhost:9014")
        # driver_path="C:/Users/courseu1/Downloads/chromedriver_win32/chromedriver.exe"
        driver_path="/Volumes/Data/chromedriver"
        chrome_options.add_argument("user-data-dir=/Users/rahulkumar/Library/Application Support/Google/Chrome/")        
        # try:
        driver = webdriver.Chrome(chrome_options=chrome_options,executable_path=driver_path)
        # except:
            # pass


        for i in driver.window_handles:
            driver.switch_to.window(i)
            driver.close()
        
        # subprocess.call('start chrome.exe -remote-debugging-port=9014 --user-data-dir-' + chrmprof,shell=True)
        # chrome_options.add_experimental_option("debuggerAddress","localhost:9014")

        try:
            driver = webdriver.Chrome(chrome_options=chrome_options,executable_path=driver_path)
        except:
            pass


        print('loading Login/Homepage')
        # driver.execute_script("window.open('https://www.google.com');")
        
        self.open_script(driver)        
        print("extesion")
        driver.switch_to.window(driver.window_handles[1])
        reg_btn_len=0
        blck_txt="No text"
        vpn_len=0
        while reg_btn_len==0:
            try:
                reg_btn_len = len(driver.find_elements_by_id('register-button'))
            except:
                reg_btn_len=0
            try:
                blck_txt = driver.find_element_by_id('main-content').text
            except:
                blck_txt="No text"
            if ("Requests to the server have been blocked by an extension" in blck_txt)==True:
                time.sleep(5)
                driver.get('chrome-extension://oofgbpoabipfcfjapgnbbjjaenockbdp/popup.html')
                blck_txt="No text"
            else:
                try:
                    vpn_len = len(driver.find_elements_by_class_name('server-item__server__server-info__server-label'))
                except:
                    vpn_len=0
                if vpn_len>0:
                    break
        if vpn_len==0:
            pass_text = "PGXUWLCZZW"
            driver.find_elements_by_class_name('login-authcode-view__authcode-container__input')[0].send_keys(pass_text)
        while vpn_len==0:
            try:
                vpn_len = len(driver.find_elements_by_class_name('server-item__server__server-info__server-label'))
            except:
                vpn_len=0
        time.sleep(5) 
        for each in driver.find_elements_by_class_name('server-item__server__server-info__server-label'):
            if each.text=="Germany":
                each.click()
                break
        conn_text="No"
        while conn_text=="No":
            try:
                conn_text = driver.find_element_by_id("connected-disconnect-button").text
            except:
                conn_text="No"
        # driver.close()
        driver.switch_to.window(driver.window_handles[0])
        prd = 241042
        self.driver = driver
        self.used_country = []
        # driver.get("https://www.saturn.de/")

        # crawling()




    def Crawling(self):
        self.Resolve_Captcha()
        df = pd.read_excel('mediamarkt_input.xlsx')
        r = df["Landing Page URL"].to_list()
        category = df["Category"].to_list()
        for links,cat in zip(r,category):
            self.crawl_link(self.driver,links,cat,links)
crawl = CrawlLinkSpider()
crawl.Crawling()
#crawl.Resolve_Captcha()
crawl.close_browser()
