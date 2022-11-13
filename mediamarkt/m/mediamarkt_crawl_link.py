import openpyxl
import os
import time
import warnings
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
# from pyvirtualdisplay import Display
import subprocess
from lxml import html
from bs4 import BeautifulSoup as BS
import pandas as pd
import time
import csv
import mysql.connector
# import MySQLdb


class CrawlLinkSpider():
    def __init__(self):
        # self.conn = MySQLdb.connect(host="75.126.130.244", user="crawler@web", passwd="data@server", db="dell_pl_crawl_latest",charset="utf8", use_unicode=True)
        self.conn = mysql.connector.connect(host="localhost", user="root", passwd="root", db="crawling",charset="utf8", use_unicode=True)
        self.cursor = self.conn.cursor()
        self.output_file_name = "PL_Crawl_Mediamarkt_DE_Dell_GPT_Laptop_Output.xlsx"
        self.driver = ""
        self.xl = open("mediamarkt_links.csv","w",encoding="utf-8")
        self.row = csv.writer(self.xl)
        self.output = pd.DataFrame()



    def Update_IP(self):
        self.driver.switch_to.window(self.driver.window_handles[1])
        try:
            country = self.driver.find_element_by_id("connected-country").text
        except:
            country = "Germany"
        
        self.used_country.append(country)
        print(self.used_country)
        try:
            self.driver.find_element_by_id("connected-disconnect-button").click()
        except:
            pass
        time.sleep(2)
        print(".....")
        print(self.driver.find_elements_by_class_name('server-item__server__server-info__server-label'))
        for each in self.driver.find_elements_by_class_name('server-item__server__server-info__server-label'):
            if(each.text not in self.used_country):
                print(each.text)
                each.click()
                break
        conn_text="No"
        while conn_text=="No":
            try:
                conn_text = self.driver.find_element_by_id("connected-disconnect-button").text
            except:
                conn_text="No"            
        self.driver.switch_to.window(self.driver.window_handles[0])
        self.driver.refresh()
        soup = BS(self.driver.page_source,'html.parser')
        if(soup.title.text == "Saturn" or soup.title.text == "MediaMarkt"):
            self.Update_IP()


    def calculate_implicit_wait_time(self,driver,wait_value,element): 
        if wait_value: 
            driver.implicitly_wait(wait_value) 
        now=time.time() 
        try: 
            myelement = driver.find_elements_by_class_name(element) 
        except: 
            pass 
        print(str(int(time.time()-now))+'--Seconds')


    def crawl_link(self,driver,url,cat,landing_page):
        driver.get(url)
        tree = html.fromstring(driver.page_source)
        if(not tree.xpath("//a[@class='Linkstyled__StyledLinkRouter-sc-1drhx1h-2 hihJjl ProductListItemstyled__StyledLink-sc-16qx04k-0 dYJAjV']/@href")):
            time.sleep(3)
        tree = html.fromstring(driver.page_source)
        # driver.implicitly_wait(10)
        self.calculate_implicit_wait_time(driver,10,"Typostyled__StyledInfoTypo-sc-1jga2g7-0")

        # tree = html.fromstring(driver.page_source)
        soup = BS(driver.page_source,'html.parser')
        if(soup.title.text == "Saturn" or soup.title.text == "MediaMarkt"):
            self.Update_IP()
        links = tree.xpath("//a[@class='Linkstyled__StyledLinkRouter-sc-1drhx1h-2 iDDAGF ProductListItemstyled__StyledLink-sc-16qx04k-0 dYJAjV']/@href")
        titles = soup.find_all("p","Typostyled__StyledInfoTypo-sc-1jga2g7-0 fuXjPV")        

        products = soup.find_all("div",attrs={"data-test":"mms-search-srp-productlist-item"})
        for product in products:
            date = time.strftime("%Y-%m-%d")
            _time = time.strftime("%H:%M:%S")            
            product_link = "https://www.mediamarkt.de"+product.find("a").get("href")
            product_title = product.find("p","Typostyled__StyledInfoTypo-sc-1jga2g7-0 fuXjPV").text
            product_code = product_link.split("-")[-1].replace(".html","")
            try:
                mrp = product.find("div","UnbrandedPricestyled__Wrapper-jah2p3-6 cscaoX").find_all("div","UnbrandedPricestyled__StyledUnbrandedPriceDisplayWrapper-jah2p3-1 jJibBP")[0].find("div","StrikeThrough__StyledStrikeThrough-sc-1uy074f-1 bBiLbg").text
                list_price = product.find("div","UnbrandedPricestyled__Wrapper-jah2p3-6 cscaoX").find_all("div","UnbrandedPricestyled__StyledUnbrandedPriceDisplayWrapper-jah2p3-1 jJibBP")[1].find("div","ToolTipstyled__StyledTooltipWrapper-sc-1rht449-0").text
            except:
                mrp = list_price = product.find("div","ToolTipstyled__StyledTooltipWrapper-sc-1rht449-0 kNYRYZ").text


            query = "insert into mediamarkt_pref(retailer,category,product_url,product_name,markdown_price,list_price,response_status,extraction_date,product_code,stock)values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
            values = ("Mediamarkt",cat,product_link,product_title,mrp,list_price,200,date,product_code,"In Stock")
            self.cursor.execute(query,values)
            self.conn.commit()

            self.cursor.execute("select * from pl_master where retailer = '{}' and product_code='{}'".format('mediamarkt',product_code))
            # self.cursor.execute("select * from pl_master where script_name = '{}' and product_code='{}' and deactivated=0".format(c["Script_Name"],c["product_id"]))
            if(not self.cursor.fetchone()):
                query = "insert into pl_master(retailer,category,DATE,TIME,LANDING_PAGE_URL,PRODUCT_TITLE,PRODUCT_LINK,PRODUCT_CODE,CUSTOMIZED_URL,NEW_URL,DEACTIVATED,REACTIVATED)VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                values = ("mediamarkt",cat,date,_time,landing_page,product_title,product_link,product_code,"","","","")
                self.cursor.execute(query,values)
                self.conn.commit()
                print(values)



        # for link,title in zip(links,titles):
        #     c = {
        #         "Date":time.strftime("%Y-%m-%d"),
        #         "Time":time.strftime("%H:%M:%S"),
        #         "Landing_Page_url":landing_page,
        #         "Product_title":title.text,
        #         "product_link":"https://www.mediamarkt.de"+link,
        #         "product_id" : link.replace(".html","").split("-")[-1],
        #         "Category":cat,
        #     }
        #     # self.row.writerow(c)
        #     print(c)
        #     query = "insert into mediamarkt_pref(retailer,category,product_url,product_name,markdown_price,list_price,response_status,extraction_date,product_code,stock)values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
        #     values = ("Mediamarkt",cat,c["product_link"],c["Product_title"],c[""])
        #     product_code = link.split("-")[-1].replace(".html","")
        #     self.cursor.execute("select * from pl_master where retailer = '{}' and product_code='{}'".format('mediamarkt',product_code))
        #     # self.cursor.execute("select * from pl_master where script_name = '{}' and product_code='{}' and deactivated=0".format(c["Script_Name"],c["product_id"]))
        #     if(not self.cursor.fetchone()):
        #         query = "insert into pl_master(retailer,category,DATE,TIME,LANDING_PAGE_URL,PRODUCT_TITLE,PRODUCT_LINK,PRODUCT_CODE,CUSTOMIZED_URL,NEW_URL,DEACTIVATED,REACTIVATED)VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
        #         values = ("mediamarkt",cat,c["Date"],c["Time"],c["Landing_Page_url"],c["Product_title"],c["product_link"],product_code,"","","","")
        #         self.cursor.execute(query,values)
        #         self.conn.commit()

        load_more = soup.find("button",attrs={"data-test":"mms-search-srp-loadmore"})
        if(load_more):
            if(driver.current_url.find("page") == -1):
                next_page_url = driver.current_url+"&page=2"
            else:
                page_num = driver.current_url.split("page=")[1]
                next_page_url = driver.current_url.replace("page={}".format(page_num),"page={}".format(int(page_num)+1))

            self.crawl_link(driver,next_page_url,cat,landing_page)
        # else:
        #     self.output.to_excel(self.output_file_name,index=False)

    def close_browser(self):
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.driver.find_element_by_id("connected-disconnect-button").click()
        time.sleep(2)
        self.driver.find_element_by_id("logout-button").click()
        for i in self.driver.window_handles:
            self.driver.switch_to.window(i)
            self.driver.close()		
        print("process completed")


    def open_script(self,driver):
        try:
            driver.execute_script("window.open('chrome-extension://oofgbpoabipfcfjapgnbbjjaenockbdp/popup.html')")
        except:
            self.open_script(driver)

    def Resolve_Captcha(self):
        print("Process Started")
        cwd = os.getcwd()
        warnings.filterwarnings("ignore", category=DeprecationWarning)
        chrome_options = Options()

        excel_input = cwd + "/Input_file.xlsx"
        wb = openpyxl.load_workbook(excel_input)
        sht_login = wb.get_sheet_by_name('Login Info')
        sht_input = wb.get_sheet_by_name('Input')
        sht_output = wb.get_sheet_by_name('Output')
        username = sht_login.cell(1,2).value
        password = sht_login.cell(2,2).value
        chrmprof = sht_login.cell(3,2).value
        subprocess.call('start chrome.exe -remote-debugging-port=9014 --user-data-dir-' + chrmprof,shell=True)
        chrome_options.add_experimental_option("debuggerAddress","localhost:9014")
        # driver_path="C:/Users/courseu1/Downloads/chromedriver_win32/chromedriver.exe"
        # driver_path = "/Volumes/Data/course5i/geckodriver"

        # prefs = {"profile.managed_default_content_settings.images": 2}
        # chrome_options.add_experimental_option("prefs", prefs)

        # driver_path="/Volumes/Data/chromedriver"
        # chrome_options.add_argument("user-data-dir=/Users/rahulkumar/Library/Application Support/Google/Chrome/")

        # try:
        driver = webdriver.Chrome(chrome_options=chrome_options)



        print('loading Login/Homepage')
        # driver.execute_script("window.open('https://www.google.com');")
        
        self.open_script(driver)
        print("extesion")
        driver.switch_to.window(driver.window_handles[1])
        reg_btn_len=0
        blck_txt="No text"
        vpn_len=0
        while reg_btn_len==0:
            try:
                reg_btn_len = len(driver.find_elements_by_id('register-button'))
            except:
                reg_btn_len=0
            try:
                blck_txt = driver.find_element_by_id('main-content').text
            except:
                blck_txt="No text"
            if ("Requests to the server have been blocked by an extension" in blck_txt)==True:
                time.sleep(5)
                driver.get('chrome-extension://oofgbpoabipfcfjapgnbbjjaenockbdp/popup.html')
                blck_txt="No text"
            else:
                try:
                    vpn_len = len(driver.find_elements_by_class_name('server-item__server__server-info__server-label'))
                except:
                    vpn_len=0
                if vpn_len>0:
                    break
        if vpn_len==0:
            pass_text = "PGXUWLCZZW"
            driver.find_elements_by_class_name('login-authcode-view__authcode-container__input')[0].send_keys(pass_text)
        while vpn_len==0:
            try:
                vpn_len = len(driver.find_elements_by_class_name('server-item__server__server-info__server-label'))
            except:
                vpn_len=0
        time.sleep(5) 
        for each in driver.find_elements_by_class_name('server-item__server__server-info__server-label'):
            if each.text=="Netherlands":
                each.click()
                break
        conn_text="No"
        while conn_text=="No":
            try:
                conn_text = driver.find_element_by_id("connected-disconnect-button").text
            except:
                conn_text="No"
        driver.switch_to.window(driver.window_handles[0])
        prd = 241042
        self.used_country = []
        self.driver = driver
    def Crawling(self):
        self.Resolve_Captcha()
        df = pd.read_excel('mediamarkt_input.xlsx')
        r = df["Landing Page URL"].to_list()
        category = df["Category"].to_list()
        for links,cat in zip(r,category):
            self.crawl_link(self.driver,links,cat,links)
crawl = CrawlLinkSpider()
crawl.Crawling()
crawl.close_browser()
