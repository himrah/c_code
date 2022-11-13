# -*- coding: utf-8 -*-
import openpyxl
import os
import requests
from datetime import datetime
import time
import warnings
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
import subprocess
from lxml import html
from bs4 import BeautifulSoup as BS
import pandas as pd
import pymysql
from lxml import html
import json
import re
import csv

class MediamarktSpiderSpider():

    def open_script(self,driver):
        try:
            driver.execute_script("window.open('chrome-extension://oofgbpoabipfcfjapgnbbjjaenockbdp/popup.html')")
        except:
            self.open_script(driver)

    def Resolve_Captcha(self):
        print("Process Started")
        cwd = os.getcwd()
        warnings.filterwarnings("ignore", category=DeprecationWarning)
        chrome_options = Options()

        excel_input = cwd + "/Input_file.xlsx"
        wb = openpyxl.load_workbook(excel_input)
        sht_login = wb.get_sheet_by_name('Login Info')
        sht_input = wb.get_sheet_by_name('Input')
        sht_output = wb.get_sheet_by_name('Output')
        username = sht_login.cell(1,2).value
        password = sht_login.cell(2,2).value
        chrmprof = sht_login.cell(3,2).value
        subprocess.call('start chrome.exe -remote-debugging-port=9014 --user-data-dir-' + chrmprof,shell=True)
        chrome_options.add_experimental_option("debuggerAddress","localhost:9014")
        prefs = {"profile.managed_default_content_settings.images": 2}

        chrome_options.add_experimental_option("prefs", prefs)
        # chrome_options.add_argument("user-data-dir=/Users/rahulkumar/Library/Application Support/Google/Chrome/")

        # driver_path="C:/Users/courseu1/Downloads/chromedriver_win32/chromedriver.exe"
        # driver_path = "C:/Users/NIXIS/Downloads/chromedriver_win32 (2)/chromedriver.exe"

        # try:
        driver = webdriver.Chrome(chrome_options=chrome_options,executable_path=driver_path)
        # except:
        #     pass
        # driver = webdriver.Chrome(executable_path="/Volumes/Data/chromedriver",chrome_options=chrome_options)


        # for i in driver.window_handles:
        #     driver.switch_to.window(i)
        #     driver.close()
        
        # subprocess.call('start chrome.exe -remote-debugging-port=9014 --user-data-dir-' + chrmprof,shell=True)
        # chrome_options.add_experimental_option("debuggerAddress","localhost:9014")

        # try:
        #     driver = webdriver.Chrome(chrome_options=chrome_options,executable_path=driver_path)
        # except:
        #     pass


        print('loading Login/Homepage')
        # driver.execute_script("window.open('https://www.google.com');")
        
        # self.open_script(driver)        
        driver.execute_script("window.open('chrome-extension://oofgbpoabipfcfjapgnbbjjaenockbdp/popup.html')")
        print("extesion")
        driver.switch_to.window(driver.window_handles[1])
        reg_btn_len=0
        blck_txt="No text"
        vpn_len=0
        while reg_btn_len==0:
            try:
                reg_btn_len = len(driver.find_elements_by_id('register-button'))
            except:
                reg_btn_len=0
            try:
                blck_txt = driver.find_element_by_id('main-content').text
            except:
                blck_txt="No text"
            if ("Requests to the server have been blocked by an extension" in blck_txt)==True:
                time.sleep(5)
                driver.get('chrome-extension://oofgbpoabipfcfjapgnbbjjaenockbdp/popup.html')
                blck_txt="No text"
            else:
                try:
                    vpn_len = len(driver.find_elements_by_class_name('server-item__server__server-info__server-label'))
                except:
                    vpn_len=0
                if vpn_len>0:
                    break
        if vpn_len==0:
            pass_text = "PGXUWLCZZW"
            driver.find_elements_by_class_name('login-authcode-view__authcode-container__input')[0].send_keys(pass_text)
        while vpn_len==0:
            try:
                vpn_len = len(driver.find_elements_by_class_name('server-item__server__server-info__server-label'))
            except:
                vpn_len=0
        time.sleep(5) 
        for each in driver.find_elements_by_class_name('server-item__server__server-info__server-label'):
            if each.text=="Germany":
                each.click()
                break
        conn_text="No"
        while conn_text=="No":
            try:
                conn_text = driver.find_element_by_id("connected-disconnect-button").text
            except:
                conn_text="No"
        driver.switch_to.window(driver.window_handles[0])
        prd = 241042
        self.driver = driver

    def close_browser(self):

        self.driver.switch_to.window(self.driver.window_handles[1])
        self.driver.find_element_by_id("connected-disconnect-button").click()
        time.sleep(2)
        self.driver.find_element_by_id("logout-button").click()
        #self.driver.close()
        #self.driver.switch_to.window(self.driver.window_handles[0])
        #self.driver.quit()
        for i in self.driver.window_handles:
           self.driver.switch_to.window(i)
           self.driver.close()
        print("process completed")

    def Crawling(self):
        # con = pymysql.connect(host="75.126.130.244", user="crawler@web", passwd="data@server", db="dell_pl_crawl_latest",charset="utf8", use_unicode=True)
        # pdp_con = pymysql.connect(host="75.126.130.244", user="crawler@web", passwd="data@server", db="dell_pdp_crawl_latest",charset="utf8", use_unicode=True)
        # pdp_cursor = pdp_con.cursor(pymysql.cursors.DictCursor)
        # cursor = con.cursor(pymysql.cursors.DictCursor)
        # cursor.execute('select * from pl_master where deactivated=0 and script_name like "%Saturn%"')
        # rows = cursor.fetchall()
        rows = open("links.txt","r",encoding="utf-8").read().split("\n")
        output = open("result.csv","w",encoding="utf-8",newline="")
        xls = csv.writer(output)
        for row in rows:
            # url = row["product_link"]
            # script_name = row["script_name"]
            url = row
            script_name = ""
            self.   driver.get(url)
            tree = html.fromstring(self.driver.page_source)
            l = []
            laptop_type = screen = os = p_series = p_model = graphics = hdd = ram = touch = resolution = optical = backlit = battery_life =  ""
            main = json.loads(tree.xpath("u//script[@type='application/ld+json']//text()")[0])
            sku = main["sku"]
            name = main["name"]
            price = "".join(re.findall("price':(.*?),",str(main)))
            mrp = "".join(re.findall('"strikePrice":(.*?),',self.driver.page_source))
            data = re.search("window.__PRELOADED_STATE__ = (.*?)};",self.driver.page_source).group(1)
            block = "".join(tree.xpath("u//section[@id='description']//text()"))
            block2 = "".join(tree.xpath("u//section[@id='features']//text()"))
            
            for i in re.findall('GraphqlProductFeature",(.*?)\}\]',data):
                l.append(json.loads("{"+i+"}]}"))
            
            for i in l:
                if(i.get("name")):
                    if(i.get("name") == "Produkttyp"):
                        laptop_type = i.get("value")
                    if(i.get("name")== "Bildschirmdiagonale (cm/Zoll)"):
                        screen = i.get("value")
                    if(i.get("name") == "Betriebssystem"):
                        os = i.get("value")
                    if(i.get("name") == "Prozessor"):
                        p_series = i.get("value")
                    if(i.get("name")=="Prozessor-Modell"):
                        p_model = i.get("value")
                    if(i.get("name")=="Grafikkarte"):
                        graphics = i.get("value")
                    if(i.get("name")== "Festplatte 1"):
                        hdd = i.get("value")
                    if(i.get("name")==u"Arbeitsspeicher-Größe"):
                        ram = i.get("value")
                    if(i.get("name") == "Touchscreen"):
                        touch = i.get("value")
                    if(i.get("name")==u"Bildschirmauflösung"):
                        resolution = i.get("value")
                    if(i.get("name") == "Laufwerkstyp"):
                        optical = i.get("value")
                    if(i.get("name") == "Tastatur"):
                        backlit = i.get("value")
                    if(i.get("name") == "Akku-Laufzeit"):
                        battery_life = i.get("value")
            # print(url)
            # query = "insert into saturn_de(script_name,competitor_list_price,competitor_markdown_price,competitor_product_id,competitor_product_name,competitor_model,competitor_product_url,competitor_laptop_type,competitor_screen_size,competitor_operating_system,competitor_processor_series,competitor_processor_model,competitor_graphics,competitor_harddrive_capacity,competitor_ram_capacity,competitor_touchscreen,competitor_display_resolution,competitor_optical_drive,competitor_keyboard_backlit,competitor_warranty,seller_name,stock,Block_data,battery,time_stamp)VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
            
            values = (script_name,mrp,price,sku,name,"",url,"",screen,os,p_series,p_model,graphics,hdd,ram,touch,resolution,optical,backlit,"","","",block+block2,battery_life,time.strftime("%Y-%m-%d %H:%M:%S"))
            xls.writerow(values)
            # pdp_cursor.execute(query,values)
            # pdp_con.commit()


            # print(["mediamarkt","laptop","lenovo",sku,"DE",laptop_type,"","",url,"","",name,mrp,price,"",p_model+" "+p_series,os,screen,ram,hdd,graphics,"","","",battery_life,resolution,"","","",block,block2])

crawl = MediamarktSpiderSpider()
crawl.Resolve_Captcha()
crawl.Crawling()
crawl.close_browser()